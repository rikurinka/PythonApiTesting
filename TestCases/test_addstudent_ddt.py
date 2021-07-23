import requests
import json
import jsonpath
import openpyxl

#data driven testing
#with multple datas in registration page
#multiplestudent from excel
#static data

def test_add_multiple_students():
    API_Url = r'http://thetestingworldapi.com/api/studentsDetails'
    f = open(r'D:\API\Addstudent.json')
    request_json = json.loads(f.read())
    #Excel code
    wk = openpyxl.load_workbook(r'D:\API\dataxl.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row

    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)

        #keyname mentioned in json request
        request_json['first_name'] = cell_first_name.value
        request_json['middle_name']= cell_mid_name.value
        request_json['last_name']= cell_last_name.value
        request_json['date_of_birth']= cell_dob.value
        response = requests.post(API_Url, request_json)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201

  #  print(request_json)














