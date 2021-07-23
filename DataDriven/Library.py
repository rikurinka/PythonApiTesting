import json
import jsonpath
import requests
import openpyxl

#Dynamicaally - generic
#two files - Library and Test case file
#how rows/colms is excel
class Common:
    #constructor
    def __init__(self, FileNamePath, SheetName):
        global wk
        global sh
        wk = openpyxl.load_workbook(FileNamePath)
        sh = wk[SheetName]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        columns = sh.max_column
        return columns

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1,c+1):
            cell = sh.cell(row=1,column=i)
            li.insert(i-1,cell.value)
        return li

    def update_request_with_data(self,rowNumber,jsonRequest,keylist):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=rowNumber, column=i)
            jsonRequest[keylist[i-1]] = cell.value
        return jsonRequest
